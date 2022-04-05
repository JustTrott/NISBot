from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import range_boundaries
import openpyxl

from config import Config

class SpreadsheetProcessor():
    def __init__(self, filename : str) -> None:
        self.book_name = filename

    def get_grade_letters(self) -> dict[str, str]:
        wb = load_workbook(filename=self.book_name, read_only=True)        
        ws = wb.active
        grades = {}
        for row in ws['A6':'A182']:
            if row[0].value is None:
                continue
            parallel = row[0].value[:-4]
            letter = row[0].value[len(parallel):-3]
            grades[parallel] = grades[parallel] + letter if parallel in grades else letter
        return grades

    def unmerge_and_fill_sheet(self) -> None:
        wb = load_workbook(filename=self.book_name, read_only=False)
        ws = wb.active
        mcr_coord_list = [mcr.coord for mcr in ws.merged_cells.ranges]
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            if min_col < 3 or min_row < 6:
                continue
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(mcr)
            for row in ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
        self.add_times_row(ws)
        wb.save(self.book_name)
    
    def add_times_row(self, ws : Worksheet):
        lessons_row = 4
        lesson_time : dict[str, tuple[str, str]] = {
            1 : ("8:15", "8:55"),
            2 : ("9:15", "9:55"),
            3 : ("10:00", "10:40"),
            4 : ("10:45", "11:25"),
            5 : ("11:30", "12:10"),
            6 : ("12:30", "13:10"),
            7 : ("13:30", "14:10"),
            8 : ("14:30", "15:10"),
            9 : ("15:25", "16:05"),
            10 : ("16:20", "17:00")
        }
        ws.insert_rows(lessons_row)
        cell : Cell
        for cell in ws[lessons_row + 1]:
            if cell.value in lesson_time:
                ws[cell.coordinate].offset(-1).value = f"{lesson_time[cell.value][0]}-{lesson_time[cell.value][1]}"
    
if __name__ == '__main__':
    sp = SpreadsheetProcessor('schedule.xlsx')
    sp.unmerge_and_fill_sheet()