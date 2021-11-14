from openpyxl import load_workbook
from openpyxl import utils
from datetime import datetime
from tabulate import tabulate

class SpreadsheetParser():
    def __init__(self, grade, weekday):
        self.grade = grade
        self.book_name = "schedule.xlsx"
        self.weekday = weekday
    def process_sheet(self, filename):
        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        for row in ws['A134':'BJ182']:
            for cell in row:
                if len(str(cell.value)) < 60:
                    continue
                target_content = str(cell.value)
                i = cell.row + 1
                j = cell.column - 1
                mergecell = type(ws[i][j])
                while type(ws[i + 1][j]) == mergecell:
                    i += 1
                if type(ws[i][cell.column]) == mergecell:
                    last_cell = ws[i][j + 1]
                else:
                    last_cell = ws[i][j]
                start_cell = cell
                try:
                    ws.unmerge_cells(f"{start_cell.coordinate}:{last_cell.coordinate}")
                except ValueError:
                    last_cell = ws[i][j]
                    ws.unmerge_cells(f"{start_cell.coordinate}:{last_cell.coordinate}")
                for k in range(start_cell.row, last_cell.row + 1, 4):
                    for l in range(start_cell.column-1, last_cell.column):
                        print(ws[k][l].coordinate)
                        ws[k][l].value = target_content
                        ws.merge_cells(f"{ws[k][l].coordinate}:{ws[k+3][l].coordinate}")
        for row in ws['C6':'BJ182']:
            for i in range(len(row) - 1):
                if type(row[i+1]) != mergecell:
                    continue
                try:
                    ws.unmerge_cells(f"{row[i].coordinate}:{row[i+1].coordinate}")
                    ws[row[i+1].coordinate].value = row[i].value
                except (ValueError, IndexError):
                    continue
        wb.save(filename)

    def parse_sheet(self, grade, weekday):
        wb = load_workbook(filename=self.book_name, read_only=True)
        ws = wb.active
        weekdays_row = ws[3]
        for cell in weekdays_row:
            if cell.value is not None and str(cell.value) == weekday:
                schedule = f"Расписание для класса {grade} на день: {str(cell.value)}\n"
                col_bounds = cell.column_letter, weekdays_row[cell.column + 10].column_letter
                break
        for row in ws:
            if str(row[0].value).startswith(grade):
                row_bounds = (row[0].row, row[0].row+3)
        schedule_items = []
        profile_subjects = []
        for row in ws[f"{col_bounds[0]}{row_bounds[0]}" : f"{col_bounds[1]}{row_bounds[1]}"]:
            data_cols = []
            for cell in row:
                if len(str(cell.value)) > 60:
                    target = f'\nПрофили {len(profile_subjects)}: {cell.value}'
                    if target not in profile_subjects:
                        profile_subjects.append(f'\nПрофили {len(profile_subjects)+1}: {cell.value}')
                        data_cols.append(f'Профили {len(profile_subjects)}')
                    else:
                        data_cols.append(f'Профили {len(profile_subjects)}')
                    continue
                data_cols.append(cell.value)
            schedule_items.append(data_cols)

        times_row = ws[f"{col_bounds[0]}4" : f"{col_bounds[1]}4"][0]
        times = []
        for cell in times_row:
            times.append(cell.value)
        schedule_items.insert(0, times)
        for i in reversed(range(len(schedule_items[1]) - 1)):
            if schedule_items[1][i] == schedule_items[1][i+1]:
                time = schedule_items[0][i]
                time = time[:time.find('-')] + schedule_items[0][i+1][schedule_items[0][i+1].find('-'):]
                schedule_items[0][i] = "p" + time
                for j in range(len(schedule_items)):
                    del schedule_items[j][i+1]
        for i in reversed(range(len(data_cols))):
            for j in range(1, len(schedule_items)):
                if schedule_items[j][i] is not None:
                    break
            else:
                for j in range(len(schedule_items)):
                    del schedule_items[j][i]
        subjects_cnt = len(schedule_items[0])
        for i in range(subjects_cnt):
            copy_schedule_items = []
            for j in range(len(schedule_items)):
                if schedule_items[j][i] is None:
                    continue
                copy_schedule_items.append([schedule_items[j][i]])
            if copy_schedule_items[0][0].startswith('p'):
                copy_schedule_items[0][0] = f'{i+1} Пара ' + copy_schedule_items[0][0][1:]
            else:
                copy_schedule_items[0][0] = f'{i+1} Урок ' + copy_schedule_items[0][0]
            schedule += tabulate(copy_schedule_items, headers='firstrow', tablefmt='pretty') + '\n'
        schedule += '\n' + '\n'.join(profile_subjects)
        return schedule

    @property
    def sheet(self):
        return self.parse_sheet(self.grade, self.weekday)


if __name__ == '__main__':
    sp = SpreadsheetParser("11D", "Пятница")
    print(sp.sheet)
    #sp.process_sheet('schedule.xlsx')
