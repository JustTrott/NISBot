from datetime import datetime

from openpyxl import load_workbook
from tabulate import tabulate

class SpreadsheetParser:
    def __init__(self, filename : str):
        self.book_name = filename

    def get_schedule(self, grade : str, weekday : str) -> str:
        if weekday == 'auto':
            weekday = self.get_current_weekday()
        self.wb = load_workbook(filename=self.book_name, read_only=True)
        ws = self.wb.active
        col_bounds = SpreadsheetParser.get_col_bounds(ws, 3, weekday)
        schedule = f"Расписание для класса {grade} на день: {weekday}\n"
        row_bounds = SpreadsheetParser.get_row_bounds(ws, 0, grade)
        schedule += SpreadsheetParser.generate_schedule(ws, row_bounds, col_bounds)
        return schedule

    def get_schedule_bulk(self, grades: 'list[str]', weekday):
        sheets: str = []
        if weekday == 'auto':
            weekday = self.get_current_weekday()
        self.wb = load_workbook(filename=self.book_name, read_only=True)
        ws = self.wb.active
        col_bounds = SpreadsheetParser.get_col_bounds(ws, 3, weekday)
        for grade in grades:
            schedule = f"Расписание для класса {grade} на день: {weekday}\n"
            row_bounds = SpreadsheetParser.get_row_bounds(ws, 0, grade)
            schedule += SpreadsheetParser.generate_schedule(ws, row_bounds, col_bounds)
            sheets.append(schedule)
        return sheets
    

    @staticmethod
    def get_row_bounds(ws, col_id, query):
        for row in ws:
            if str(row[col_id].value).startswith(query):
                row_bounds = (row[col_id].row, row[col_id].row+3)
        return row_bounds


    @staticmethod
    def get_col_bounds(ws, row_id, query):
        row = ws[row_id]
        for cell in row:
            if cell.value is not None and str(cell.value) == query:
                col_bounds = cell.column_letter, row[cell.column + 8].column_letter
                break
        return col_bounds


    @staticmethod
    def generate_schedule(ws, row_bounds, col_bounds):
        schedule_items = []
        profile_subjects = []
        schedule = ""
        for row in ws[f"{col_bounds[0]}{row_bounds[0]}" : f"{col_bounds[1]}{row_bounds[1]}"]:
            data_cols = []
            for cell in row:
                if len(str(cell.value)) < 60:
                    data_cols.append(cell.value)
                    continue
                target = f'{cell.value}'
                if target not in profile_subjects:
                    profile_subjects.append(target)
                    data_cols.append(f'Профили{len(profile_subjects)}')
                else:
                    data_cols.append(None)
            schedule_items.append(data_cols)
        times_row = ws[f"{col_bounds[0]}4" : f"{col_bounds[1]}4"][0]
        times = []
        for cell in times_row:
            times.append(cell.value)
        schedule_items.insert(0, times)
        for i in reversed(range(len(schedule_items[1]) - 1)):
            if schedule_items[1][i] != schedule_items[1][i+1]:
                continue
            time = schedule_items[0][i]
            time = time[:time.find('-')] + schedule_items[0][i+1][schedule_items[0][i+1].find('-'):]
            schedule_items[0][i] = "p" + time
            for j in range(len(schedule_items)):
                del schedule_items[j][i+1]
        for i in reversed(range(len(data_cols))):
            for j in range(1, len(schedule_items)):
                if schedule_items[j][i] is not None:
                    break
            else:
                for j in range(len(schedule_items)):
                    del schedule_items[j][i]
        subjects_cnt = len(schedule_items[0])
        for i in range(subjects_cnt):
            copy_schedule_items = []
            for j in range(len(schedule_items)):
                if schedule_items[j][i] is None:
                    continue
                copy_schedule_items.append([schedule_items[j][i]])
            if copy_schedule_items[0][0].startswith('p'):
                copy_schedule_items[0][0] = f'{i+1} Пара ' + copy_schedule_items[0][0][1:]
            else:
                copy_schedule_items[0][0] = f'{i+1} Урок ' + copy_schedule_items[0][0]
            schedule += tabulate(copy_schedule_items, headers='firstrow', tablefmt='pretty') + '\n'
        schedule += '\n' + '\n\n'.join([f'Профили{index+1}: {subject}' for index, subject in enumerate(profile_subjects)])
        return schedule


    def get_current_weekday(self) -> str:
        weekdays = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
        if datetime.utcnow().hour >= 10:
            day = datetime.utcnow().weekday() + 1
        else:
            day = datetime.utcnow().weekday()
        if day > 4:
            day = 0
        return weekdays[day]


if __name__ == '__main__':
    sp = SpreadsheetParser('schedule.xlsx')
    print(sp.get_schedule('11D', 'Понедельник'))
